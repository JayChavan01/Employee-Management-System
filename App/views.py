from django.shortcuts import redirect, render
from .forms import EmployeeForm
from .models import Employee

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# ------------------ Home Page ------------------
def Home(request):
    return render(request, 'index.html')

# ------------------ Employee Registration ------------------
def Register_Employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('details')  # Redirect to employee details page after registration
    else:
        form = EmployeeForm()

    context = {'form': form}
    return render(request, 'register.html', context)

# ------------------ Employee Details ------------------
def Employee_Details(request):
    data = Employee.objects.all()
    context = {'Data': data}
    return render(request, 'details.html', context)

# ------------------ Delete Employee ------------------
def Delete_Record(request, id):
    employee = Employee.objects.get(pk=id)
    employee.delete()
    return redirect('/')  # Redirect back to details page after deletion

# ------------------ Update Employee ------------------
def Update_Record(request, id):
    data = Employee.objects.get(pk=id)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=data)
        if form.is_valid():
            form.save()
            return redirect('/')  # Redirect to details page after update
    else:
        form = EmployeeForm(instance=data)

    context = {'form': form}
    return render(request, 'update.html', context)

# ------------------ Export Data ------------------
def Export_Employees(request):
    employees = Employee.objects.all()
    
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Data"
    
    # Define column headers
    headers = [
        "ID", "First Name", "Last Name", "Email", 
        "Position", "Mobile Number", "Join Date"
    ]
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write employee data
    for row_num, employee in enumerate(employees, 2):
        ws.cell(row=row_num, column=1, value=employee.id)
        ws.cell(row=row_num, column=2, value=employee.first_name)
        ws.cell(row=row_num, column=3, value=employee.last_name)
        ws.cell(row=row_num, column=4, value=employee.email)
        ws.cell(row=row_num, column=5, value=employee.position)
        ws.cell(row=row_num, column=6, value=employee.mobile_number)
        ws.cell(row=row_num, column=7, value=employee.join_date.strftime('%Y-%m-%d'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Employee Data.xlsx'
    wb.save(response)
    
    return response